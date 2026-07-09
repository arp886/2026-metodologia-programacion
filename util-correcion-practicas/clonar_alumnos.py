#!/usr/bin/env python3
import argparse
import json
import os
import re
import subprocess
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def remove_accents(value):
    normalized = unicodedata.normalize("NFD", value)
    return "".join(char for char in normalized if unicodedata.category(char) != "Mn")


def format_name_part(value):
    value = remove_accents(value)
    value = re.sub(r"[^A-Za-z ]", " ", value)
    return "".join(part.capitalize() for part in value.split())


def split_full_name(full_name):
    parts = str(full_name).strip().split()
    if len(parts) < 2:
        raise ValueError(f"Nombre incompleto: {full_name!r}")

    surnames = parts[:2]
    names = parts[2:] or parts[1:]
    return surnames, names


def repo_name_from_student(full_name, suffix):
    surnames, names = split_full_name(full_name)
    return f"{format_name_part(' '.join(surnames))}{format_name_part(' '.join(names))}{suffix}"


def parse_grade(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def excel_column_index(column):
    if isinstance(column, int):
        return column
    return column_index_from_string(str(column).strip())


def approved_students(workbook_path, sheet_name, name_column, grade_column, min_grade):
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"No existe la hoja {sheet_name!r}. Hojas disponibles: {available}")

    sheet = workbook[sheet_name]
    name_index = excel_column_index(name_column) - 1
    grade_index = excel_column_index(grade_column) - 1
    for row in sheet.iter_rows(min_row=1):
        full_name = row[name_index].value if len(row) > name_index else None
        grade = parse_grade(row[grade_index].value if len(row) > grade_index else None)
        if full_name and grade is not None and grade > min_grade:
            yield str(full_name).strip(), grade


def authenticated_url(base_url, username, token, repo_name):
    base_url = base_url.rstrip("/")
    if token:
        return base_url.replace("https://", f"https://{username}:{token}@", 1) + f"/{repo_name}.git"
    return f"{base_url}/{repo_name}.git"


def public_url(base_url, repo_name):
    return f"{base_url.rstrip('/')}/{repo_name}"


def clone_with_git(repo_url, repo_dir):
    return subprocess.run(
        ["git", "clone", repo_url, str(repo_dir)],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )


def clone_repo(repo_name, destination, github, token, dry_run=False, label=None, no_auth=False):
    repo_dir = destination / repo_name
    visible_url = public_url(github["organization_unit"], repo_name)

    if repo_dir.exists():
        subject = label or repo_name
        print(f"Saltando {subject}: ya existe {repo_dir}")
        return

    subject = f"{label}: " if label else ""
    print(f"Clonando {subject}{visible_url}")
    if not dry_run:
        urls = []
        if token and not no_auth:
            urls.append(authenticated_url(github["organization_unit"], github["username"], token, repo_name))
        urls.append(f"{visible_url}.git")

        last_output = ""
        for index, repo_url in enumerate(urls):
            result = clone_with_git(repo_url, repo_dir)
            if result.returncode == 0:
                return
            last_output = result.stdout
            if index == 0 and len(urls) > 1:
                print("Fallo el clonado autenticado; reintentando sin token...")

        print(f"No se pudo clonar {visible_url}")
        if last_output:
            print(last_output.strip())
        raise SystemExit(1)


def main():
    parser = argparse.ArgumentParser(description="Clona repositorios de alumnos aprobados de MP2026.")
    parser.add_argument("--config", default="config.json", help="Ruta al fichero de configuracion")
    parser.add_argument("--dry-run", action="store_true", help="Muestra lo que haria sin clonar")
    parser.add_argument("--repo", help="Clona un repositorio concreto por nombre, por ejemplo ZapataRojasMiguelMP2026")
    parser.add_argument("--no-auth", action="store_true", help="Clona usando la URL publica, sin token")
    args = parser.parse_args()

    root = Path(__file__).resolve().parent
    config_path = (root / args.config).resolve()
    config = json.loads(config_path.read_text(encoding="utf-8"))

    github = config["github"]
    excel = config["excel"]
    clone = config["clone"]

    token = os.environ.get("GITHUB_TOKEN", github.get("token", ""))
    destination = (root / clone.get("destination", "alumnos")).resolve()
    destination.mkdir(parents=True, exist_ok=True)

    workbook_path = root / excel["file"]
    suffix = clone.get("suffix", "MP2026")

    if args.repo:
        clone_repo(args.repo, destination, github, token, dry_run=args.dry_run, no_auth=args.no_auth)
        return

    for full_name, grade in approved_students(
        workbook_path,
        excel["sheet"],
        excel.get("student_name_column", "B"),
        excel.get("approved_grade_column", "J"),
        excel.get("approved_min_grade", 5),
    ):
        repo_name = repo_name_from_student(full_name, suffix)
        clone_repo(repo_name, destination, github, token, dry_run=args.dry_run, label=f"{full_name} ({grade})", no_auth=args.no_auth)


if __name__ == "__main__":
    main()
