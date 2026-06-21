#!/usr/bin/env python3
import argparse
import json
import re
import shutil
import subprocess
import tempfile
import xml.etree.ElementTree as ET
from contextlib import contextmanager
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

from clonar_alumnos import repo_name_from_student


def run(command, cwd, timeout=None):
    return subprocess.run(
        command,
        cwd=cwd,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        timeout=timeout,
    )


def git_commit_for_session(repo, session, extra_days):
    start = date.fromisoformat(session["start"])
    end = date.fromisoformat(session["end"]) + timedelta(days=extra_days)
    since = f"{start.isoformat()} 00:00:00"
    until = f"{end.isoformat()} 23:59:59"

    commits = []
    for code in session.get("codes", []):
        result = run(
            ["git", "log", "--all", "--format=%H|%cs|%s", f"--since={since}", f"--until={until}", f"--grep={code}"],
            repo,
        )
        commits.extend(line for line in result.stdout.splitlines() if line.strip())

    path_result = run(
        ["git", "log", "--all", "--format=%H|%cs|%s", f"--since={since}", f"--until={until}", "--", f"src/main/java/org/mp/{session['path']}", f"src/test/java/org/mp/{session['path']}", f"Test/org/mp/{session['path']}"],
        repo,
    )
    commits.extend(line for line in path_result.stdout.splitlines() if line.strip())

    unique = list(dict.fromkeys(commits))
    return bool(unique), unique[:5]


def test_root(repo):
    return repo / "src" / "test" / "java"


def test_classes_for_session(repo, session_path):
    classes = []
    root = test_root(repo)
    session_root = root / "org" / "mp" / session_path
    if not session_root.exists():
        return []
    for file_path in sorted(session_root.rglob("*Test*.java")):
        relative = file_path.relative_to(root).with_suffix("")
        classes.append(".".join(relative.parts))
    return list(dict.fromkeys(classes))


def test_files_for_session(repo, session_path):
    session_root = test_root(repo) / "org" / "mp" / session_path
    if not session_root.exists():
        return []
    return sorted(session_root.rglob("*Test*.java"))


def copy_teacher_session_tests(teacher_repo, repo, session_path):
    teacher_session = test_root(teacher_repo) / "org" / "mp" / session_path
    student_session = test_root(repo) / "org" / "mp" / session_path
    if not teacher_session.exists():
        return

    rewrites = package_rewrites_for_repo(repo)
    for source in sorted(teacher_session.rglob("*")):
        if source.is_dir():
            continue
        if source.suffix == ".java" and not source.name.endswith("Test.java"):
            continue
        relative = source.relative_to(teacher_session)
        destination = student_session / relative
        destination.parent.mkdir(parents=True, exist_ok=True)
        if source.suffix == ".java":
            text = source.read_text(encoding="utf-8", errors="ignore")
            for old, new in rewrites.items():
                text = text.replace(old, new)
            destination.write_text(text, encoding="utf-8")
        else:
            shutil.copy2(source, destination)


def copy_teacher_test_file(teacher_repo, repo, source):
    rewrites = package_rewrites_for_repo(repo)
    root = test_root(teacher_repo)
    relative = source.relative_to(root)
    destination = test_root(repo) / relative
    destination.parent.mkdir(parents=True, exist_ok=True)
    text = source.read_text(encoding="utf-8", errors="ignore")
    for old, new in rewrites.items():
        text = text.replace(old, new)
    destination.write_text(text, encoding="utf-8")
    return ".".join(destination.relative_to(test_root(repo)).with_suffix("").parts)


def has_java(repo, relative_path):
    return (repo / "src" / "main" / "java" / relative_path).exists()


def package_rewrites_for_repo(repo):
    rewrites = {}
    if has_java(repo, Path("org/mp/sesion01/ListaDeReproduccion.java")):
        rewrites["org.mp.sesion01.listareproduccion"] = "org.mp.sesion01"
        rewrites["package org.mp.sesion01.listareproduccion;"] = "package org.mp.sesion01;"
    if has_java(repo, Path("org/mp/sesion03/Par.java")):
        rewrites["org.mp.sesion03.pares"] = "org.mp.sesion03"
        rewrites["package org.mp.sesion03.pares;"] = "package org.mp.sesion03;"
    return rewrites


def count_test_methods(files):
    total = 0
    for file_path in files:
        text = file_path.read_text(encoding="utf-8", errors="ignore")
        total += len(re.findall(r"@(?:org\.junit\.jupiter\.api\.)?Test\b", text))
        total += len(re.findall(r"@(?:org\.junit\.jupiter\.params\.)?ParameterizedTest\b", text))
        total += len(re.findall(r"@(?:org\.junit\.jupiter\.api\.)?RepeatedTest\b", text))
    return total


@contextmanager
def temporary_teacher_tests(repo, teacher_repo, session_path):
    student_test_root = test_root(repo)
    backup_root = repo / "target" / "evaluador-backup-student-tests"

    if student_test_root.exists():
        backup_root.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(student_test_root), str(backup_root))
    copy_teacher_session_tests(teacher_repo, repo, session_path)

    try:
        yield
    finally:
        if student_test_root.exists():
            shutil.rmtree(student_test_root)
        if backup_root.exists():
            student_test_root.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(backup_root), str(student_test_root))


@contextmanager
def temporary_teacher_test_file(repo, teacher_repo, source):
    student_test_root = test_root(repo)
    backup_root = repo / "target" / "evaluador-backup-student-tests"

    if student_test_root.exists():
        backup_root.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(student_test_root), str(backup_root))
    class_name = copy_teacher_test_file(teacher_repo, repo, source)

    try:
        yield class_name
    finally:
        if student_test_root.exists():
            shutil.rmtree(student_test_root)
        if backup_root.exists():
            student_test_root.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(backup_root), str(student_test_root))


@contextmanager
def temporarily_hide_misplaced_tests(repo):
    main_root = repo / "src" / "main" / "java"
    hidden_root = repo / "target" / "evaluador-hidden-main-tests"
    moved = []
    if main_root.exists():
        for file_path in sorted(main_root.rglob("*Test*.java")):
            relative = file_path.relative_to(main_root)
            destination = hidden_root / relative
            destination.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(file_path), str(destination))
            moved.append((file_path, destination))

    try:
        yield
    finally:
        for original, hidden in reversed(moved):
            original.parent.mkdir(parents=True, exist_ok=True)
            if hidden.exists():
                shutil.move(str(hidden), str(original))
        if hidden_root.exists():
            shutil.rmtree(hidden_root)


@contextmanager
def temporarily_hide_other_sessions(repo, current_session_path):
    main_session_root = repo / "src" / "main" / "java" / "org" / "mp"
    hidden_root = repo / "target" / "evaluador-hidden-other-sessions"
    moved = []

    if main_session_root.exists():
        for directory in sorted(main_session_root.iterdir()):
            if not directory.is_dir():
                continue
            if not re.fullmatch(r"sesion\d+", directory.name):
                continue
            if directory.name == current_session_path:
                continue
            destination = hidden_root / directory.name
            destination.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(directory), str(destination))
            moved.append((directory, destination))

    try:
        yield
    finally:
        for original, hidden in reversed(moved):
            original.parent.mkdir(parents=True, exist_ok=True)
            if hidden.exists():
                shutil.move(str(hidden), str(original))
        if hidden_root.exists():
            shutil.rmtree(hidden_root)


def compilation_error_sources(output, repo):
    sources = []
    main_root = (repo / "src" / "main" / "java").resolve()
    for match in re.finditer(r"\[ERROR\]\s+(.+?\.java):\[", output):
        source = Path(match.group(1)).resolve()
        try:
            source.relative_to(main_root)
        except ValueError:
            continue
        if source not in sources:
            sources.append(source)
    return sources


def hide_compile_error_source(repo, source):
    hidden_root = repo / "target" / "evaluador-hidden-compile-errors"
    relative = source.relative_to(repo.resolve())
    destination = hidden_root / relative
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(source), str(destination))
    return source, destination


def restore_hidden_compile_sources(repo, moved):
    for original, hidden in reversed(moved):
        original.parent.mkdir(parents=True, exist_ok=True)
        if hidden.exists():
            shutil.move(str(hidden), str(original))
    hidden_root = repo / "target" / "evaluador-hidden-compile-errors"
    if hidden_root.exists():
        shutil.rmtree(hidden_root)


def restore_interrupted_evaluator_state(repo):
    main_session_root = repo / "src" / "main" / "java" / "org" / "mp"
    hidden_sessions = repo / "target" / "evaluador-hidden-other-sessions"
    if hidden_sessions.exists():
        main_session_root.mkdir(parents=True, exist_ok=True)
        for directory in sorted(hidden_sessions.iterdir()):
            destination = main_session_root / directory.name
            if not destination.exists():
                shutil.move(str(directory), str(destination))
        shutil.rmtree(hidden_sessions)

    hidden_sources = repo / "target" / "evaluador-hidden-compile-errors"
    if hidden_sources.exists():
        for source in sorted(hidden_sources.rglob("*")):
            if not source.is_file():
                continue
            relative = source.relative_to(hidden_sources)
            destination = repo / relative
            if not destination.exists():
                destination.parent.mkdir(parents=True, exist_ok=True)
                shutil.move(str(source), str(destination))
        shutil.rmtree(hidden_sources)

    backup_tests = repo / "target" / "evaluador-backup-student-tests"
    student_test_root = test_root(repo)
    if backup_tests.exists():
        if student_test_root.exists():
            shutil.rmtree(student_test_root)
        student_test_root.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(backup_tests), str(student_test_root))


def parse_surefire_reports(repo, session_path):
    reports_dir = repo / "target" / "surefire-reports"
    totals = {"tests": 0, "failures": 0, "errors": 0, "skipped": 0}
    if not reports_dir.exists():
        return totals

    marker = f".org.mp.{session_path}."
    for report in reports_dir.glob("TEST-*.xml"):
        try:
            root = ET.parse(report).getroot()
        except ET.ParseError:
            continue
        class_name = root.attrib.get("name", "")
        if marker not in f".{class_name}.":
            continue
        for key in totals:
            totals[key] += int(float(root.attrib.get(key, 0)))
    return totals


@contextmanager
def temporary_evaluation_repo(repo):
    with tempfile.TemporaryDirectory(prefix=f"evaluador-{repo.name}-") as temp_dir:
        work_repo = Path(temp_dir) / repo.name
        shutil.copytree(
            repo,
            work_repo,
            ignore=shutil.ignore_patterns(".git", "target"),
        )
        yield work_repo


def evaluate_session(repo, work_repo, teacher_repo, session, timeout, extra_days):
    has_commit, commit_samples = git_commit_for_session(repo, session, extra_days)
    teacher_files = test_files_for_session(teacher_repo, session["path"])
    classes = test_classes_for_session(teacher_repo, session["path"])
    expected_total = count_test_methods(teacher_files)
    reports_dir = work_repo / "target" / "surefire-reports"
    if reports_dir.exists():
        shutil.rmtree(reports_dir)

    if not classes:
        return {
            "commit_ok": has_commit,
            "commit_samples": commit_samples,
            "test_classes": 0,
            "tests_total": expected_total,
            "tests_passed": 0,
            "failures": 0,
            "errors": 0,
            "skipped": 0,
            "status": "sin_tests",
        }

    totals = {"tests": 0, "failures": 0, "errors": 0, "skipped": 0}
    failed_commands = 0
    for teacher_file in teacher_files:
        if reports_dir.exists():
            shutil.rmtree(reports_dir)
        hidden_compile_sources = []
        with temporary_teacher_test_file(work_repo, teacher_repo, teacher_file) as class_name, temporarily_hide_misplaced_tests(work_repo), temporarily_hide_other_sessions(work_repo, session["path"]):
            try:
                for _ in range(10):
                    if reports_dir.exists():
                        shutil.rmtree(reports_dir)
                    result = run(
                        ["mvn", "-q", "-DfailIfNoTests=false", f"-Dtest={class_name}", "test"],
                        work_repo,
                        timeout=timeout,
                    )
                    if result.returncode == 0:
                        break

                    sources = compilation_error_sources(result.stdout, work_repo)
                    new_sources = [source for source in sources if source.exists()]
                    if not new_sources:
                        break
                    for source in new_sources:
                        hidden_compile_sources.append(hide_compile_error_source(work_repo, source))
                else:
                    result = run(
                        ["mvn", "-q", "-DfailIfNoTests=false", f"-Dtest={class_name}", "test"],
                        work_repo,
                        timeout=timeout,
                    )
            finally:
                restore_hidden_compile_sources(work_repo, hidden_compile_sources)
        if result.returncode != 0:
            failed_commands += 1
        class_totals = parse_surefire_reports(work_repo, session["path"])
        for key in totals:
            totals[key] += class_totals[key]

    executed_total = totals["tests"]
    total = max(expected_total, executed_total)
    passed = executed_total - totals["failures"] - totals["errors"] - totals["skipped"]
    status = "ok" if failed_commands == 0 else "tests_fallidos_o_error"
    if totals["tests"] == 0 and failed_commands:
        status = "error_compilacion_o_maven"

    return {
        "commit_ok": has_commit,
        "commit_samples": commit_samples,
        "test_classes": len(classes),
        "tests_total": total,
        "tests_passed": passed,
        "failures": totals["failures"],
        "errors": totals["errors"],
        "skipped": totals["skipped"],
        "status": status,
    }


def commit_value(result):
    return "correcto" if result and result["commit_ok"] else "NO"


def test_value(result):
    if not result:
        return "0 / 0"
    return f"{result['tests_passed']} / {result['tests_total']}"


def write_results_workbook(root, config, results_by_repo):
    excel_config = config["excel"]
    evaluation = config["evaluation"]
    sessions = evaluation["sessions"]
    workbook_path = root / excel_config["file"]
    workbook = load_workbook(workbook_path)
    values_workbook = load_workbook(workbook_path, data_only=True)
    if excel_config["sheet"] not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"No existe la hoja {excel_config['sheet']!r}. Hojas disponibles: {available}")

    sheet = workbook[excel_config["sheet"]]
    values_sheet = values_workbook[excel_config["sheet"]]
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=column)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = values_sheet.cell(row=row, column=column).value

    start_column = sheet.max_column + 1
    for index, session in enumerate(sessions):
        column = start_column + index * 2
        session_label = session["name"].replace("Session", "sesion")
        sheet.cell(row=1, column=column, value=f"{session_label} commit")
        sheet.cell(row=1, column=column + 1, value=f"{session_label} test")

    suffix = config["clone"].get("suffix", "MP2026")
    for row in range(2, sheet.max_row + 1):
        full_name = values_sheet.cell(row=row, column=2).value or sheet.cell(row=row, column=2).value
        if not full_name:
            continue
        try:
            repo_name = repo_name_from_student(full_name, suffix)
        except ValueError:
            continue
        repo_results = results_by_repo.get(repo_name, {})
        for index, session in enumerate(sessions):
            result = repo_results.get(session["name"])
            column = start_column + index * 2
            sheet.cell(row=row, column=column, value=commit_value(result))
            sheet.cell(row=row, column=column + 1, value=test_value(result))

    output_path = root / evaluation["results_xlsx"]
    workbook.save(output_path)
    return output_path


def evaluate_repo(repo, teacher_repo, evaluation):
    print(f"Evaluando {repo.name}", flush=True)
    restore_interrupted_evaluator_state(repo)
    repo_results = {}
    details = []
    with temporary_evaluation_repo(repo) as work_repo:
        for session in evaluation["sessions"]:
            data = evaluate_session(
                repo,
                work_repo,
                teacher_repo,
                session,
                evaluation.get("java_timeout_seconds", 90),
                evaluation.get("commit_extra_days", 7),
            )
            row = {
                "repo": repo.name,
                "session": session["name"],
                "commit_ok": data["commit_ok"],
                "test_classes": data["test_classes"],
                "tests_passed": data["tests_passed"],
                "tests_total": data["tests_total"],
                "failures": data["failures"],
                "errors": data["errors"],
                "skipped": data["skipped"],
                "status": data["status"],
            }
            repo_results[session["name"]] = row
            details.append({**row, "commit_samples": data["commit_samples"]})
    return repo.name, repo_results, details


def main():
    parser = argparse.ArgumentParser(description="Evalua commits y tests por sesion en repos Java Maven.")
    parser.add_argument("--config", default="config.json")
    parser.add_argument("--repo", help="Evalua solo un repositorio por nombre")
    args = parser.parse_args()

    root = Path(__file__).resolve().parent
    config = json.loads((root / args.config).read_text(encoding="utf-8"))
    evaluation = config["evaluation"]
    students_dir = root / config["clone"]["destination"]
    teacher_repo = (root / evaluation["teacher_repo"]).resolve()
    if not (teacher_repo / "src" / "test" / "java").exists():
        raise ValueError(f"No existe el directorio de tests del profesor: {teacher_repo}")

    repos = sorted(path for path in students_dir.iterdir() if (path / ".git").exists())
    if args.repo:
        repos = [path for path in repos if path.name == args.repo]

    max_workers = int(evaluation.get("max_workers", 1))
    results_by_repo = {}
    details_by_repo = {}
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(evaluate_repo, repo, teacher_repo, evaluation): repo for repo in repos}
        for future in as_completed(futures):
            repo = futures[future]
            try:
                repo_name, repo_results, repo_details = future.result()
            except Exception as exc:
                raise RuntimeError(f"Error evaluando {repo.name}: {exc}") from exc
            results_by_repo[repo_name] = repo_results
            details_by_repo[repo_name] = repo_details

    details = []
    for repo in repos:
        details.extend(details_by_repo.get(repo.name, []))

    teacher_totals = {
        session["name"]: count_test_methods(test_files_for_session(teacher_repo, session["path"]))
        for session in evaluation["sessions"]
    }
    json_path = root / evaluation["results_json"]
    json_path.write_text(json.dumps(details, ensure_ascii=False, indent=2), encoding="utf-8")
    xlsx_path = write_results_workbook(root, config, results_by_repo)
    print(f"Resultados escritos en {json_path.name} y {xlsx_path.name}")
    print("Totales tests profesor: " + ", ".join(f"{name}={total}" for name, total in teacher_totals.items()))


if __name__ == "__main__":
    main()
